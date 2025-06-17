"""Function to write your merged file into your database files"""

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries, get_column_letter


def load_data(new_df, workbook_path, sheet_name, table_name, batch_name):
    """
    Add new data to a database stored as Table in a given excel workbook.

    params:
        new_df (pd.DataFrame): DataFrame containing the new data to be added to the database.
        workbook_path (str): Path to the Excel workbook where the database is stored.
        sheet_name (str): Name of the sheet within the workbook where the database table is located.
        table_name (str): Name of the table within the sheet where the data is stored.
        batch_name (str): Name or identifier for the batch associated with the new data.
    """

    print("Loading data into the database...")  # for debug
    # PREPARATION OF THE NEW DATA
    ## Add the missing columns to the DataFrame
    new_df["BATCH NAME"] = batch_name
    new_df["CONCENTRATION\n[ug/mL]"] = ""
    print(new_df.head())

    ## Reorder the columns to match the database
    new_df = new_df[
        [
            "ACQUISITION DATE",
            "BATCH NAME",
            "CHROMATOGRAM NAME",
            "CHANNEL",
            "ANALITE",
            "AREA",
            "CONCENTRATION\n[ug/mL]",
        ]
    ]
    # print(new_df.head())

    # LOAD THE DATABASE
    # Load the existing database from the specified sheet
    database = pd.read_excel(workbook_path, sheet_name)
    print(database)

    # Change date format (avoid time)
    # database["FECHA"] = pd.to_datetime(database["FECHA"]).dt.strftime("%d-%m-%Y")
    # print(database)

    # Remove completely empty rows
    rows = [r for r in dataframe_to_rows(new_df, index=False, header=False) if any(r)]

    # --- LOAD EXISTING WORKBOOK AND TABLE ---
    wb = load_workbook(workbook_path)  # Load the existing workbook
    ws = wb[sheet_name]  # Find the existing sheet in the workbook
    table = ws.tables[table_name]  # Find the existing table in the worksheet

    # Determine table range
    table_range = table.ref  # e.g., "A1:G25"
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    start_row = max_row + 1  # start appending just after the current table
    end_row = start_row + len(rows) - 1

    # Copy the last formula in the CONCENTRATION column (if any)
    formula_col = max_col  # 7 for column G
    last_formula_cell = ws.cell(row=max_row, column=formula_col)
    formula_to_copy = (
        last_formula_cell.value
        if isinstance(last_formula_cell.value, str)
        and last_formula_cell.value.startswith("=")
        else None
    )

    # --- APPEND NEW DATA ---
    for r_idx, row in enumerate(rows, start=start_row):  # row loop
        for c_idx, value in enumerate(row, start=1):  # column loop
            # create cell object
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # For the last column (formula), apply formula if available
            if c_idx == formula_col and formula_to_copy:
                cell.value = formula_to_copy

    # --- UPDATE TABLE RANGE SAFELY ---
    end_col_letter = get_column_letter(max_col)
    table.ref = f"A1:{end_col_letter}{end_row}"

    # --- SAVE WORKBOOK ---
    wb.save(workbook_path)
    print("¡Data loaded successfully into the database!")
