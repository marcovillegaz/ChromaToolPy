"""In ChromNAV you can extract your peak info in an excel file by channels. For
further analysis, this script extract the important info (ex. peak area, peak name,
etc.) and merged them in a single excel file, where each sheet correspond to a 
single analyte, this facilitates the calibration curve analysis."""

from post_processing.merged import *
from post_processing.data_base import *

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils.dataframe import dataframe_to_rows

channels = ["CH1", "CH9", "CH10", "CH11", "CH12", "CH13", "CH14", "CH15"]
filepath = r"C:\Users\marco\Escritorio\Henry - Column(1)_out"
extract_cols = ["Chromatogram Name", "CH", "Peak Name", "Area"]

# Specify the values for the date and experiment columns
date_value = "2024-07-27"  # Replace with the specified date
experiment_value = "experiment_name"  # Replace with the specified experiment name
database_file = r"C:\Users\marco\OneDrive - usach.cl\GWR Analysis\Hojas de calculo\HG_samples - copia.xlsx"

# Merge single channel data
appended_df = merge_files(filepath, channels, extract_cols)
print(appended_df.head())

# Change columns name.
appended_df.rename(
    columns={
        "Chromatogram Name": "CHROMATOGRAM NAME",
        "CH": "CHANNEL",
        "Peak Name": "ANALITE",
        "Area": "AREA",
    },
    inplace=True,
)
print(appended_df.head())

# Add the missing columns to the DataFrame
appended_df["FECHA"] = date_value
appended_df["EXPERIMENTAL RUN"] = experiment_value
appended_df["CONCENTRATION\n[ug/mL]"] = ""
print(appended_df.head())

# Reorder the columns to match the database
appended_df = appended_df[
    [
        "FECHA",
        "EXPERIMENTAL RUN",
        "CHROMATOGRAM NAME",
        "CHANNEL",
        "ANALITE",
        "AREA",
        "CONCENTRATION\n[ug/mL]",
    ]
]
print(appended_df.head())

# Load the existing database from the specified sheet
database = pd.read_excel(database_file, sheet_name="TABLE")
database["FECHA"] = pd.to_datetime(database["FECHA"]).dt.date
print(database)

# Load the existing workbook and sheet
wb = load_workbook(database_file)
ws = wb["TABLE"]
table = ws.tables["ScreeningTable"]  # Find the existing table in the worksheet

# Get the last row of the table
last_row = table.ref.split(":")[-1].split("$")[-1]

# Append the new data to the existing database
updated_database = pd.concat([database, appended_df], ignore_index=True)
# print(updated_database)

# Convert the updated DataFrame back to rows
rows = dataframe_to_rows(updated_database, index=False, header=False)

# Find the last row of the table
last_row = ws.max_row

# Append the rows to the existing table
for r_idx, row in enumerate(rows, start=int(last_row) + 1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Apply the formula to the concentration column if it's the last column
        if c_idx == len(row):
            cell.value = ws.cell(
                row=int(last_row), column=c_idx
            ).value  # Copy the formula from the last row

# Update the table reference to include new rows
table.ref = f"A1:G{ws.max_row}"

# Save the workbook
wb.save(database_file)

print(
    "Data appended successfully while preserving the table format and applying the formula."
)
