"""Function to write your merged file into your database files"""

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def load_data(new_df, workbook_path, sheet_name, table_name, date, experiment_name):
    """
    Add new data to a database stored as Table in a given excel workbook.

    params:
        new_df (pd.DataFrame): DataFrame containing the new data to be added to the database.
        workbook_path (str): Path to the Excel workbook where the database is stored.
        sheet_name (str): Name of the sheet within the workbook where the database table is located.
        table_name (str): Name of the table within the sheet where the data is stored.
        date (str): Date associated with the new data entry, typically used for logging or record-keeping.
        experiment_name (str): Name or identifier for the experiment associated with the new data.
    """

    # PREPARATION OF THE NEW DATA
    ## Add the missing columns to the DataFrame
    new_df["FECHA"] = date
    new_df["EXPERIMENTAL RUN"] = experiment_name
    new_df["CONCENTRATION\n[ug/mL]"] = ""
    # print(new_df.head())

    ## Reorder the columns to match the database
    new_df = new_df[
        [
            "FECHA",
            "EXPERIMENTAL RUN",
            "CHROMATOGRAM NAME",
            "CHANNEL",
            "ANALITE",
            "AREA",
            "CONCENTRATION\n[ug/mL]",
        ]
    ]
    # print(new_df.head())

    # LOAD THE DATABASE
    # Load the existing database from the specified sheet
    database = pd.read_excel(workbook_path, sheet_name)
    # Change date format (avoid time)
    database["FECHA"] = pd.to_datetime(database["FECHA"]).dt.strftime("%d-%m-%Y")
    # print(database)

    wb = load_workbook(workbook_path)  # Load the existing workbook
    ws = wb[sheet_name]  # Find the existing sheet in the workbook
    table = ws.tables[table_name]  # Find the existing table in the worksheet

    # Convert the updated DataFrame back to rows
    rows = dataframe_to_rows(new_df, index=False, header=False)

    # Find the last row of the table
    last_row = ws.max_row

    # Append the rows to the existing table
    for r_idx, row in enumerate(rows, start=int(last_row) + 1):  # row loop
        for c_idx, value in enumerate(row, start=1):  # column loop
            # create cell object
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Apply the formula to the concentration column if it's the last column
            if c_idx == len(row):
                # Copy the formula from the last row
                cell.value = ws.cell(row=int(last_row), column=c_idx).value

    # Update the table reference to include new rows
    table.ref = f"A1:G{ws.max_row}"
    # Save the workbook
    wb.save(workbook_path)

    print("¡Data loaded successfully into the database!")
